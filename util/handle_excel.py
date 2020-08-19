# encoding: utf-8
"""
@author: yanNa
@software: PyCharm
@file: handle_excel.py
@time: 2020/8/14 10:02
"""
import os

import openpyxl


def get_file_path():
    file_path = os.path.dirname(os.getcwd()) + r'\config\case_data.xlsx'
    return file_path


class HandleExcel(object):
    def __init__(self):
        self.index = 0

    def __index__(self, index):
        self.index = index

    def load_excel(self):
        open_excel = openpyxl.load_workbook(get_file_path())
        return open_excel

    def get_sheet_value(self, index=None):
        '''加载所有sheet的内容'''
        sheet_names = self.load_excel().sheetnames
        if index is None:
            index = 0
        self.index = index
        data = self.load_excel()[sheet_names[self.index]]
        return data

    def get_cell_value(self, row, cols):
        data = self.get_sheet_value(self.index).cell(row, cols).value
        return data

    def get_rows(self):
        '''获取行数'''
        row = self.get_sheet_value(self.index).max_row
        return row

    def get_row_values(self, row):
        '''获取某一行内容'''
        row_list = []
        '''获取对应行的内容'''
        for i in self.get_sheet_value(self.index)[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self, row, col, value):
        '''写入数据'''
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row, col, value)
        wb.save(get_file_path())

    def get_colums_value(self, key=None):
        '''获取某一列数据'''
        columns_list = []
        if key is None:
            key = 'A'
        columns_list_data = self.get_sheet_value()[key]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list

    def get_row_number(self, content, key=None):
        '''获取行号'''
        num = 1
        cols_data = self.get_colums_value(key)
        for col_data in cols_data:
            if str(col_data) == content:
                return num
            num = num + 1
        return num

'''单例'''
excel_data = HandleExcel()

if __name__ == '__main__':
    handle = HandleExcel()
    print(handle.get_cell_value(1, 2))
    print(handle.get_rows())
    print(handle.get_row_values(1))
    print(handle.get_row_number('1'))

